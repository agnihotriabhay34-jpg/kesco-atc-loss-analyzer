"""Generate downloadable AT&C analysis reports (xlsx, csv, pdf) from analysis results.

Two levels:
  - "summary": headline KPIs + division table
  - "detailed": everything (KPIs, divisions, feeders, outstanding, causes, payment modes)

Each builder returns raw bytes so the Flask layer can stream them as a download.
"""

import io
import csv
import datetime as _dt


def _now():
    return _dt.datetime.now().strftime("%Y-%m-%d %H:%M")


def _summary_rows(data):
    """Flatten the summary block into label/value rows."""
    s = data["summary"][0]
    return [
        ("Report period", s.get("month", "") if s.get("month") != "Data snapshot" else "Not specified"),
        ("Consumers analysed", s.get("consumers", 0)),
        ("Billing efficiency (%)", s.get("billing_eff_pct")),
        ("Collection efficiency (%)", s.get("collection_eff_pct")),
        ("AT&C loss (%)", s.get("atc_loss_pct")),
        ("Energy input matched (MU)", s.get("input_matched_mu")),
        ("Theft suspects", s.get("theft_suspects", 0)),
        ("Total outstanding (Rs cr)", s.get("outstanding_cr")),
    ]


# ---------------- CSV ----------------
def build_csv(data, level="summary"):
    out = io.StringIO()
    w = csv.writer(out)
    s = data["summary"][0]
    month = s.get("month", "")
    title_suffix = f" — {month}" if month and month != "Data snapshot" else ""
    w.writerow([f"KESCO AT&C Analysis Report{title_suffix}"])
    w.writerow([f"Generated: {_now()}"])
    w.writerow([])

    w.writerow(["SUMMARY"])
    for label, val in _summary_rows(data):
        w.writerow([label, val])
    w.writerow([])

    m = data["months"][month] if month in data.get("months", {}) else None

    if level == "detailed" and m:
        w.writerow(["DIVISIONS"])
        w.writerow(["Division", "Consumers", "Billing Eff %", "Collection Eff %", "AT&C %", "Theft"])
        for d in m.get("divisions", []):
            w.writerow([d.get("div_name"), d.get("consumers"), d.get("billing_eff"),
                        d.get("collection_eff"), d.get("atc_loss"), d.get("theft_suspects")])
        w.writerow([])

        w.writerow(["FEEDERS"])
        w.writerow(["Feeder", "Billing Eff %", "Theft Suspects", "Input Under-recorded"])
        for f in m.get("feeders", []):
            w.writerow([f.get("feeder"), f.get("billing_eff"),
                        f.get("theft_suspects"), f.get("input_under_recorded")])
        w.writerow([])

        w.writerow(["OUTSTANDING BY DIVISION"])
        w.writerow(["Division", "Dues (cr)", "Credits (cr)", "Net (cr)"])
        for o in m.get("outstanding", {}).get("by_division", []):
            w.writerow([o.get("name"), o.get("dues_cr"), o.get("credits_cr"), o.get("outstanding_cr")])
        w.writerow([])

        w.writerow(["LOSS CAUSES BY DIVISION"])
        w.writerow(["Division", "Theft Suspects", "No Meter", "Estimated Bills"])
        for c in m.get("causes_by_division", []):
            w.writerow([c.get("div_name"), c.get("theft_suspects"),
                        c.get("no_meter"), c.get("estimated_bills")])

    return out.getvalue().encode("utf-8")


# ---------------- XLSX ----------------
def build_xlsx(data, level="summary"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    s = data["summary"][0]
    month = s.get("month", "")
    m = data["months"].get(month) if month in data.get("months", {}) else None

    wb = Workbook()
    HEAD = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    HFILL = PatternFill("solid", fgColor="1F4E78")
    TITLE = Font(bold=True, size=14, name="Arial", color="1F4E78")
    BOLD = Font(bold=True, name="Arial")
    thin = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HEAD; cell.fill = HFILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 3, 45)

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    xl_suffix = f" — {month}" if month and month != "Data snapshot" else ""
    ws["A1"] = f"KESCO AT&C Analysis{xl_suffix}"; ws["A1"].font = TITLE
    ws["A2"] = f"Generated: {_now()}"; ws["A2"].font = Font(italic=True, color="808080")
    r = 4
    ws.cell(row=r, column=1, value="Metric"); ws.cell(row=r, column=2, value="Value")
    style_header(ws, r, 2); r += 1
    for label, val in _summary_rows(data):
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=2, value=val)
        r += 1
    autosize(ws)

    if level == "detailed" and m:
        # Divisions
        wd = wb.create_sheet("Divisions")
        cols = ["Division", "Consumers", "Billing Eff %", "Collection Eff %", "AT&C %", "Theft"]
        wd.append(cols); style_header(wd, 1, len(cols))
        for d in m.get("divisions", []):
            wd.append([d.get("div_name"), d.get("consumers"), d.get("billing_eff"),
                       d.get("collection_eff"), d.get("atc_loss"), d.get("theft_suspects")])
        autosize(wd)

        # Feeders
        wf = wb.create_sheet("Feeders")
        cols = ["Feeder", "Billing Eff %", "Theft Suspects", "Input Under-recorded"]
        wf.append(cols); style_header(wf, 1, len(cols))
        for f in m.get("feeders", []):
            wf.append([f.get("feeder"), f.get("billing_eff"),
                       f.get("theft_suspects"), "Yes" if f.get("input_under_recorded") else "No"])
        autosize(wf)

        # Outstanding
        wo = wb.create_sheet("Outstanding")
        cols = ["Division", "Dues (Rs cr)", "Credits (Rs cr)", "Net (Rs cr)"]
        wo.append(cols); style_header(wo, 1, len(cols))
        for o in m.get("outstanding", {}).get("by_division", []):
            wo.append([o.get("name"), o.get("dues_cr"), o.get("credits_cr"), o.get("outstanding_cr")])
        autosize(wo)

        # Causes by division
        wc = wb.create_sheet("Loss Causes")
        cols = ["Division", "Theft Suspects", "No Meter", "Estimated Bills"]
        wc.append(cols); style_header(wc, 1, len(cols))
        for c in m.get("causes_by_division", []):
            wc.append([c.get("div_name"), c.get("theft_suspects"),
                       c.get("no_meter"), c.get("estimated_bills")])
        autosize(wc)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------- PDF ----------------
def build_pdf(data, level="summary"):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    s = data["summary"][0]
    month = s.get("month", "")
    m = data["months"].get(month) if month in data.get("months", {}) else None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm)
    styles = getSampleStyleSheet()
    NAVY = colors.HexColor("#1F4E78")
    ORANGE = colors.HexColor("#E87722")
    title = ParagraphStyle("t", parent=styles["Title"], textColor=NAVY, fontSize=18)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=NAVY)
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=8, textColor=colors.grey)
    story = []

    story.append(Paragraph("KESCO AT&C Analysis Report", title))
    period_txt = f"Period: {month} &nbsp;|&nbsp; " if month and month != "Data snapshot" else ""
    story.append(Paragraph(f"{period_txt}Generated: {_now()}", small))
    story.append(Spacer(1, 10))

    def tbl(header, rows, widths=None):
        data_rows = [header] + rows
        t = Table(data_rows, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    # KPI summary table
    story.append(Paragraph("Summary", h2))
    kpi_rows = [[str(l), str(v)] for l, v in _summary_rows(data)]
    story.append(tbl(["Metric", "Value"], kpi_rows, widths=[90 * mm, 60 * mm]))
    story.append(Spacer(1, 12))

    if level == "detailed" and m:
        divs = m.get("divisions", [])
        if divs:
            story.append(Paragraph("AT&C Loss by Division", h2))
            rows = [[d.get("div_name"), d.get("consumers"), d.get("billing_eff"),
                     d.get("collection_eff"), d.get("atc_loss"), d.get("theft_suspects")] for d in divs]
            rows = [[str(x) for x in r] for r in rows]
            story.append(tbl(["Division", "Consumers", "Bill %", "Coll %", "AT&C %", "Theft"], rows))
            story.append(Spacer(1, 12))

        out = m.get("outstanding", {}).get("by_division", [])
        if out:
            story.append(Paragraph("Outstanding by Division", h2))
            rows = [[o.get("name"), o.get("dues_cr"), o.get("credits_cr"), o.get("outstanding_cr")] for o in out]
            rows = [[str(x) for x in r] for r in rows]
            story.append(tbl(["Division", "Dues (cr)", "Credits (cr)", "Net (cr)"], rows))
            story.append(Spacer(1, 12))

        feeders = m.get("feeders", [])
        if feeders:
            story.append(PageBreak())
            story.append(Paragraph("Feeders (by billing efficiency)", h2))
            rows = [[f.get("feeder"), f.get("billing_eff"), f.get("theft_suspects")] for f in feeders]
            rows = [[str(x) for x in r] for r in rows]
            story.append(tbl(["Feeder", "Billing Eff %", "Theft"], rows,
                             widths=[90 * mm, 40 * mm, 30 * mm]))

    doc.build(story)
    return buf.getvalue()


# ---------------- dispatcher ----------------
def build_report(data, fmt="xlsx", level="summary"):
    fmt = (fmt or "xlsx").lower()
    level = (level or "summary").lower()
    if fmt == "csv":
        return build_csv(data, level), "text/csv", "csv"
    if fmt == "pdf":
        return build_pdf(data, level), "application/pdf", "pdf"
    return (build_xlsx(data, level),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx")
